import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
from ttkthemes import ThemedTk
from crypto_utils import derive_key, encrypt_password, decrypt_password
from excel_utils import add_password_to_excel
from openpyxl import load_workbook
import pyotp
import qrcode
import os
from PIL import Image, ImageTk


class LoginWindow:
    def __init__(self, parent, login_function):
        self.parent = parent
        self.login_function = login_function
        self.setup_window()

    def setup_window(self):
        self.top = ctk.CTkToplevel(self.parent)
        self.top.geometry("535x396")  # Adjust the window size to fit the entire application
        self.top.title("Login Page")

        # Load background image
        BG_IMAGE_PATH = 'bck.jpg'  # Make sure the path is correct relative to where your script is run
        bg_image = Image.open(BG_IMAGE_PATH)
        bg_photo = ImageTk.PhotoImage(bg_image)

        # Create a label for the background image and place it on the left side
        bg_label = tk.Label(self.top, image=bg_photo)
        bg_label.image = bg_photo  # Keep a reference to prevent garbage collection
        bg_label.place(x=0, y=0, width=350, height=396)


        # Adjust the position of the login frame to the right
        login_frame_width = self.top.winfo_width() - 350  # Remaining width after the bg image
        login_frame_height = self.top.winfo_height()  # Full height
        login_frame = ctk.CTkFrame(self.top, width=login_frame_width, height=login_frame_height)
        login_frame.place(x=350, y=0)  # Place it right after the image

        # Welcome label
        welcome_label = ctk.CTkLabel(login_frame, text="Welcome Back!")  # Removed text_font parameter
        welcome_label.pack(pady=12)

        # Password Entry
        password_entry = ctk.CTkEntry(login_frame, placeholder_text="Password:", show="*")
        password_entry.pack(pady=10, padx=20, fill='x')

        # Login button
        login_button = ctk.CTkButton(login_frame, text="Login", command=lambda: self.on_login(password_entry.get()))
        login_button.pack(pady=20)

    def on_login(self, master_password):
        if self.login_function(master_password):
            self.top.destroy()
        else:
            ctk.CTkMessageBox.show_error("Login Failed", "The provided master password is incorrect.")



class PasswordManagerGUI:
    def __init__(self, root, master_password):
        self.root = root
        self.master_password = master_password
        self.key = derive_key(master_password)
        ctk.set_appearance_mode("dark")
        self.root.title("Password Manager by Aplik v1.0")
        self.root.configure(bg='#333333')  # Example hex color for dark theme
        self.style = ttk.Style()
        self.style.configure("TLabel", background="#333333", foreground="white")  # Example configuration
        self.create_widgets()

    def create_widgets(self):
        labels_texts = ["Site Name:", "Username:", "Password:", "Notes:"]
        self.entries = []
        for i, text in enumerate(labels_texts):
            label = ttk.Label(self.root, text=text)
            label.grid(column=0, row=i, padx=10, pady=10, sticky="W")
            entry = ttk.Entry(self.root)
            entry.grid(column=1, row=i, padx=10, pady=10, sticky="EW")
            self.entries.append(entry)

        add_button = ctk.CTkButton(self.root, text="Add Password", command=self.add_password)
        add_button.grid(column=0, row=4, columnspan=2, padx=10, pady=20)

        add_button = ctk.CTkButton(self.root, text="View Passwords", command=self.view_passwords)
        add_button.grid(column=0, row=5, columnspan=2, padx=10, pady=15)

        self.root.grid_columnconfigure(1, weight=1)



    def add_password(self):
        site, username, password, notes = (entry.get() for entry in self.entries)
        encrypted_password, salt = encrypt_password(password,
                                                    self.master_password)
        add_password_to_excel(site, username, encrypted_password, salt, notes)

        for entry in self.entries:
            entry.delete(0, tk.END)
        print(f"Password for {site} added successfully.")

    def view_passwords(self):
        filename = "passwords.xlsx"
        wb = load_workbook(filename)
        sheet = wb.active

        new_window = tk.Toplevel(self.root)
        new_window.title("View Stored Passwords")
        new_window.configure(bg='#333333')

        # Sorting controls inside new_window
        sort_order = ttk.Combobox(new_window, values=["Ascending", "Descending"], state="readonly")
        sort_order.grid(row=0, column=0, padx=10, pady=5)
        sort_order.set("Ascending")

        sort_button = ctk.CTkButton(new_window, text="Sort",
                                    command=lambda: self.fill_passwords(sheet, new_window, "Platform",
                                                                        sort_order.get().lower()))
        sort_button.grid(row=0, column=1, padx=10, pady=5)

        self.fill_passwords(sheet, new_window)

    def fill_passwords(self, sheet, window, sort_by=None, sort_order='ascending'):
        # Clear the existing entries in the window
        for widget in window.grid_slaves():
            if int(widget.grid_info()["row"]) > 0:  # Assume row 0 is the sorting controls
                widget.grid_forget()

        # Fetch passwords starting from the first row as there are no headers
        passwords = [
            (row[0], row[1], decrypt_password(row[2], self.master_password, row[3]), row[4] or "")
            for row in sheet.iter_rows(min_row=1, values_only=True)  # Start from the first row
        ]

        if sort_by:
            sort_index = ["Platform", "Username", "Notes"].index(sort_by)
            passwords.sort(key=lambda x: x[sort_index], reverse=(sort_order == 'descending'))

        # Display all passwords
        for i, (site, username, decrypted_password, notes) in enumerate(passwords,
                                                                        start=1):  # Start at 1 for the first data row
            tk.Label(window, text=site, bg='#333333', fg='white').grid(row=i, column=0, padx=10, pady=5, sticky='w')
            tk.Label(window, text=username, bg='#333333', fg='white').grid(row=i, column=1, padx=10, pady=5, sticky='w')
            tk.Label(window, text=decrypted_password, bg='#333333', fg='white').grid(row=i, column=2, padx=10, pady=5,
                                                                                     sticky='w')
            tk.Label(window, text=notes, bg='#333333', fg='white').grid(row=i, column=3, padx=10, pady=5, sticky='w')

            delete_button = tk.Button(window, text="X", command=lambda rn=i: self.confirm_delete(sheet, window, rn))
            delete_button.grid(row=i, column=4, padx=10, pady=5, sticky='w')


    def confirm_delete(self, sheet, window, excel_row):
        confirm = tk.messagebox.askyesno("Confirm Deleting", "This will delete the credentials, proceed?")
        if confirm:
            # Delete the row from the sheet
            sheet.delete_rows(excel_row)

            # Save the workbook
            sheet.parent.save("passwords.xlsx")

            # Refresh the displayed passwords
            self.fill_passwords(sheet, window)



    def apply_sort(self, sheet, window, sort_by, sort_order):
        self.fill_passwords(sheet, window, sort_by=sort_by, sort_order=sort_order)

